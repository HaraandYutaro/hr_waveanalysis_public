#! /usr/bin/python3
"""
SEG2 file format reader and writer

This module now serves as a thin backward-compatible facade. The
:class:`SEG2Reader` implementation has been physically extracted into
:mod:`src.converter.seg2_reader` as part of the SEG-2 reading-layer
separation slice. Existing imports such as
``from src.converter.seg2 import SEG2Reader`` continue to resolve via
the re-export below, so no call-site changes are required.

The legacy ``if __name__ == "__main__":`` block at the bottom of this
file is preserved as-is (Excel datasheet driven SEG-2 → npz packaging)
and is intentionally left untouched in this slice.
"""

import glob
import os

import numpy as np
import openpyxl as xl

# Backward-compatible re-export. Keep this import so that
# `from src.converter.seg2 import SEG2Reader` continues to work
# without any call-site changes.
import sys

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from src.converter.seg2_reader import SEG2Reader  # noqa: F401


# Placeholder for SEG2Writer class (to be implemented if needed)
if __name__ == "__main__":
    """
    SEG2 file writer (placeholder for future implementation)
    """
    ###fig title と　savenameは別にする

    print("start")
    # DATfileのリストを作成
    DAT_dir = "testcodes/data/amenomiya20260629/dats/"
    sg2_list = glob.glob(DAT_dir + "*.DAT")
    dir_npz = "testcodes/data/amenomiya20260629/npzs/"
    if not os.path.exists(dir_npz):
        os.makedirs(dir_npz)

    ##excelfileから値を取得する
    xlfile = xl.load_workbook(
        "testcodes/data/amenomiya20260629/datasheet/datasheet_amenomiya.xlsx"
    )
    st1 = xlfile["Sheet1"]
    # データの取得___type:str
    dataNo_from = 100
    dataNo_to = 101

    unit = "v"  ### ジオフォンは加速度計なので unit == 'v'

    finish_name = "GEO_" + str(st1.cell(dataNo_to, 1).value)

    for i in range(dataNo_from, dataNo_to):
        row = i
        datanum = st1.cell(row, 1).value  # 番号
        shot = st1.cell(row, 2).value  # 起振方法
        condition = st1.cell(row, 3).value  # 側線条件
        excdistance = st1.cell(
            row, 4
        ).value  # 受信点とセンサー1との距離(センサーが並ぶ方向が正)
        source_x = float(st1.cell(row, 5).value)  # 起振点位置x
        sensor1_x = float(st1.cell(row, 6).value)  # x(sensor1)
        interval = float(st1.cell(row, 7).value)  # 間隔
        Num_sensor = int(st1.cell(row, 8).value)  # x(sensor24)

        xchs = st1.cell(row, 9).value   # x方向のch
        ychs = st1.cell(row, 10).value  # y方向のch
        zchs = st1.cell(row, 11).value  # z方向のch

        x_reverse = st1.cell(
            row, 12
        ).value  # x方向を逆にして計測しているかどうか:1->+-が逆になっている
        y_reverse = st1.cell(row, 13).value
        z_reverse = st1.cell(row, 14).value

        distance = np.ones(Num_sensor, dtype=np.float64) * 99
        for k in range(Num_sensor):
            d_k = np.float64(st1.cell(row, 15 + k).value)

            if np.isnan(d_k):
                # 何も書かれていないとき->  intervalとsensor1_xから距離を算出
                distance[k] = sensor1_x + interval * k
            else:
                # 特殊条件としてchxの距離が入力されている->そのchの位置を書き換える
                distance[k] = d_k

        # データを選択
        name = "GEO_" + str(datanum)
        if len(str(datanum)) == 1:
            name = "GEO_000" + str(datanum)
        if len(str(datanum)) == 2:
            name = "GEO_00" + str(datanum)
        if len(str(datanum)) == 3:
            name = "GEO_0" + str(datanum)
        dataname = name + ".DAT"
        if len(glob.glob(DAT_dir + dataname)) == 0:
            print(name + " is not in directory")
            continue
        DATfile = glob.glob(DAT_dir + dataname)[0]  # listで出てくるので[0]必須
        data = SEG2Reader(DATfile).get_all_numpy_array().T
        fs = SEG2Reader(DATfile).get_frequency()

        # +-を考慮して、各軸の値を取得
        if xchs != None:  # x方向記述あり
            xchs_from = int(xchs.split("-")[0])  # x方向のch
            xchs_to = int(xchs.split("-")[-1])  # x方向のch
            x = data[xchs_from - 1 : xchs_to, :]
            x = -x if x_reverse == 1 else x
        if ychs != None:  # y方向記述あり
            ychs_from = int(ychs.split("-")[0])
            ychs_to = int(ychs.split("-")[-1])
            y = data[ychs_from - 1 : ychs_to, :]
            y = -y if y_reverse == 1 else y
        if zchs != None:  # z方向記述あり
            zchs_from = int(zchs.split("-")[0])
            zchs_to = int(zchs.split("-")[-1])
            z = data[zchs_from - 1 : zchs_to, :]
            z = -z if z_reverse == 1 else z

        # dir_npz + nameという名前で保存
        # 今まで作ったすべての要素をattributeで保存
        # dir_npz + nameという名前で保存
        np.savez(
            dir_npz + name,
            x=x if xchs != None else None,
            y=y if ychs != None else None,
            z=z if zchs != None else None,
            data=data,
            fs=fs,
            shot=shot,
            condition=condition,
            excdistance=excdistance,
            source_x=source_x,
            sensor1_x=sensor1_x,
            interval=interval,
            Num_sensor=Num_sensor,
            distance=distance,
            unit=unit,
        )
        print(name)
