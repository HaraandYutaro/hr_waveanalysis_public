#! /usr/bin/python3
"""
Datasheet parser for the canonical seismic-survey Excel workbook.

This module is the dedicated **datasheet parsing layer** introduced as
Slice 2 of the SEG-2 / JSON sidecar input refactor (see
``private/docs/ai_refactor.md`` §3 and §6). Its sole responsibility is
to read the canonical workbook
``sample_data/datasheet/datasheet_example.xlsx`` and turn one observation
row into a normalized in-memory record.

This module:

- never modifies the Excel workbook (opened with ``read_only=True``,
  no ``save()`` is ever called)
- never reads or writes ``.sg2`` files
- never produces JSON sidecars or NPZ snapshots
- never touches ``SingleProcesser`` or any plotting code

The canonical workbook is expected to contain three sheets, and may
optionally contain a fourth:

- ``Sheet1``      survey-/acquisition-level metadata, keyed by ``No``
- ``z_distance``  per-channel z-distance arrays, keyed by the same ``No``
- ``x_distance``  per-channel x-distance arrays, keyed by the same ``No``
- ``distance``    (optional) per-channel ground-distance arrays, keyed
  by the same ``No``. When this sheet is present and contains a row
  for the requested observation id, its values become the canonical
  source for :attr:`NormalizedMetadata.distance`; otherwise that field
  is left as ``None`` and downstream consumers fall back to the
  ``sensor1_x + i*interval`` derivation.

Public entry points:

- :func:`load_datasheet`  — open the workbook, scan all rows, return a
  fully-loaded :class:`DatasheetIndex` (the workbook handle is closed
  before returning).
- :func:`get_record`      — pull one normalized
  :class:`NormalizedMetadata` record out of the index by observation id.

Design boundaries for this slice (intentionally narrow):

- Column binding is by **header name only** (never positional indices).
- Schema mismatches fail loudly with human-readable :class:`ValueError`
  messages that include sheet name, row number, and observation id.
- Cross-checks against ``SEG2Reader.get_max_ch()`` are explicitly
  out-of-scope for this slice (deferred to the dataset-loader slice).
- The legacy ``Sheet1.x_chN`` columns are intentionally not consumed
  here; the canonical x-distance source is the ``x_distance`` sheet.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Mapping, Optional, Tuple, Union

import openpyxl as xl


# ---------------------------------------------------------------------------
# Canonical workbook schema (frozen for this phase).
# ---------------------------------------------------------------------------

_SHEET_META: str = "Sheet1"
_SHEET_Z: str = "z_distance"
_SHEET_X: str = "x_distance"
# Optional fourth sheet. Intentionally NOT underscored, per project spec.
_SHEET_DISTANCE: str = "distance"

# Required Sheet1 headers (exact strings as they appear in the workbook).
_HDR_NO: str = "No"
_HDR_SHOT_METHOD: str = "起振方法"
_HDR_CONDITION: str = "側線条件"
_HDR_EXCDISTANCE: str = "起振点とセンサー1との距離(センサーの並ぶ方向が正)"
_HDR_SOURCE_X: str = "起振位置x"
_HDR_SENSOR1_X: str = "x(sensor1)"
_HDR_INTERVAL: str = "間隔"
_HDR_NUM_SENSOR: str = "センサー数"
_HDR_X_CHS: str = "X方向のch"
_HDR_Y_CHS: str = "Y方向のch"
_HDR_Z_CHS: str = "Z方向のch"
_HDR_X_REVERSE: str = "X-反転"
_HDR_Y_REVERSE: str = "Y-反転"
_HDR_Z_REVERSE: str = "Z-反転"

_REQUIRED_META_HEADERS: Tuple[str, ...] = (
    _HDR_NO,
    _HDR_SHOT_METHOD,
    _HDR_CONDITION,
    _HDR_EXCDISTANCE,
    _HDR_SOURCE_X,
    _HDR_SENSOR1_X,
    _HDR_INTERVAL,
    _HDR_NUM_SENSOR,
    _HDR_X_CHS,
    _HDR_Y_CHS,
    _HDR_Z_CHS,
    _HDR_X_REVERSE,
    _HDR_Y_REVERSE,
    _HDR_Z_REVERSE,
)

# Required headers on z_distance / x_distance sheets (the join column).
_REQUIRED_DIST_HEADERS: Tuple[str, ...] = (_HDR_NO,)

# Per-channel column prefix on z_distance / x_distance sheets ("ch1", "ch2", ...).
_DIST_CHANNEL_PREFIX: str = "ch"


# ---------------------------------------------------------------------------
# Public data types.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class NormalizedMetadata:
    """One observation's normalized metadata, parsed from the workbook.

    Channel arrays (``z_distance`` / ``x_distance``) are exposed as
    ``tuple[float, ...]`` so the record stays hashable and immutable.
    They are indexed in **total-channel** space: a survey may record
    several components per physical sensor (e.g. 24 sensors x 2
    components Y+Z = 48 channels), so their length equals
    ``num_sensor * n_components`` (``n_components >= 1``) after
    validation in :func:`get_record`. For single-component sheets this
    reduces to ``num_sensor``.

    The optional ``distance`` field carries the per-channel ground-
    distance array sourced from the optional ``distance`` sheet of the
    workbook. It is ``None`` when the sheet is absent or does not
    contain a row for the requested observation id; consumers must
    treat ``None`` as "fall back to the existing derivation"
    (``sensor1_x + i*interval``). When present, its length matches the
    same total-channel length as ``z_distance`` / ``x_distance`` after
    validation in :func:`get_record`.
    """

    obs_id: Any
    shot_method: Any
    condition: Any
    excdistance: Any
    source_x: float
    sensor1_x: float
    interval: float
    num_sensor: int
    x_chs_range: Any
    y_chs_range: Any
    z_chs_range: Any
    x_reverse: Any
    y_reverse: Any
    z_reverse: Any
    z_distance: Tuple[float, ...]
    x_distance: Tuple[float, ...]
    distance: Optional[Tuple[float, ...]] = None


@dataclass(frozen=True)
class DatasheetIndex:
    """In-memory index built by :func:`load_datasheet`.

    All sheet contents needed to answer :func:`get_record` queries are
    pre-extracted into plain Python dicts, so the workbook handle is
    closed before this object is returned and no file I/O happens at
    lookup time.
    """

    path: Path
    # Observation key (str-normalized) -> {header_name: raw cell value}
    meta_records: Mapping[str, Mapping[str, Any]]
    # Observation key (str-normalized) -> per-channel z values (ch1, ch2, ...)
    z_records: Mapping[str, Tuple[float, ...]]
    # Observation key (str-normalized) -> per-channel x values (ch1, ch2, ...)
    x_records: Mapping[str, Tuple[float, ...]]
    # Optional per-channel ground-distance values, keyed the same way.
    # Empty when the workbook lacks a 'distance' sheet entirely.
    distance_records: Mapping[str, Tuple[float, ...]]
    # Source row numbers (1-based, including the header row) for error messages.
    meta_row_numbers: Mapping[str, int]
    z_row_numbers: Mapping[str, int]
    x_row_numbers: Mapping[str, int]
    distance_row_numbers: Mapping[str, int]


# ---------------------------------------------------------------------------
# Public API.
# ---------------------------------------------------------------------------


def load_datasheet(path: Union[str, Path]) -> DatasheetIndex:
    """Open the canonical workbook read-only and build a :class:`DatasheetIndex`.

    The Excel workbook is opened with ``read_only=True, data_only=True``
    and closed before this function returns. The returned index holds
    only plain Python dicts, so no file handle is retained.

    Raises
    ------
    FileNotFoundError
        If ``path`` does not exist.
    ValueError
        If a required sheet is missing, a required header is missing,
        a duplicate observation id is found, or any other schema
        violation is detected.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Datasheet not found: {path}")

    wb = xl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in (_SHEET_META, _SHEET_Z, _SHEET_X):
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Datasheet '{path}' is missing required sheet "
                    f"'{sheet_name}'. Found sheets: {wb.sheetnames}."
                )

        meta_records, meta_rows = _index_meta_sheet(
            wb[_SHEET_META], path, _REQUIRED_META_HEADERS
        )
        z_records, z_rows = _index_dist_sheet(
            wb[_SHEET_Z], path, _REQUIRED_DIST_HEADERS
        )
        x_records, x_rows = _index_dist_sheet(
            wb[_SHEET_X], path, _REQUIRED_DIST_HEADERS
        )
        # The 'distance' sheet is optional. When absent we expose empty
        # maps so downstream lookups simply find no override entry.
        if _SHEET_DISTANCE in wb.sheetnames:
            distance_records, distance_rows = _index_dist_sheet(
                wb[_SHEET_DISTANCE], path, _REQUIRED_DIST_HEADERS
            )
        else:
            distance_records, distance_rows = {}, {}
    finally:
        wb.close()

    return DatasheetIndex(
        path=path,
        meta_records=meta_records,
        z_records=z_records,
        x_records=x_records,
        distance_records=distance_records,
        meta_row_numbers=meta_rows,
        z_row_numbers=z_rows,
        x_row_numbers=x_rows,
        distance_row_numbers=distance_rows,
    )


def get_record(
    index: DatasheetIndex, obs_id: Union[int, str]
) -> NormalizedMetadata:
    """Return the :class:`NormalizedMetadata` for one observation id.

    The lookup key is str-normalized so ``891`` and ``"891"`` are
    treated as the same record. The per-channel arrays in
    ``z_distance`` and ``x_distance`` are indexed in total-channel
    space, so their length must be a positive integer multiple of
    ``num_sensor`` from ``Sheet1.センサー数`` (``num_sensor * n_components``)
    and must agree with each other; violations raise a
    :class:`ValueError` with sheet/row/observation context.

    Raises
    ------
    ValueError
        If the observation id is missing from any of the three sheets,
        if ``センサー数`` is missing or non-integer, or if the channel
        array lengths are not a consistent positive multiple of
        ``センサー数``.
    """
    key = _obs_key(obs_id)
    if key is None:
        raise ValueError(
            f"Observation id must be a non-empty value, got {obs_id!r}."
        )

    if key not in index.meta_records:
        raise ValueError(
            f"Observation id '{obs_id}' not found in sheet "
            f"'{_SHEET_META}' of datasheet '{index.path}'."
        )
    if key not in index.z_records:
        raise ValueError(
            f"Observation id '{obs_id}' not found in sheet "
            f"'{_SHEET_Z}' of datasheet '{index.path}' "
            f"(present in '{_SHEET_META}' at row "
            f"{index.meta_row_numbers[key]})."
        )
    if key not in index.x_records:
        raise ValueError(
            f"Observation id '{obs_id}' not found in sheet "
            f"'{_SHEET_X}' of datasheet '{index.path}' "
            f"(present in '{_SHEET_META}' at row "
            f"{index.meta_row_numbers[key]})."
        )

    meta_row_num = index.meta_row_numbers[key]
    z_row_num = index.z_row_numbers[key]
    x_row_num = index.x_row_numbers[key]
    meta = index.meta_records[key]

    raw_num_sensor = meta.get(_HDR_NUM_SENSOR)
    num_sensor = _coerce_num_sensor(
        raw_num_sensor,
        obs_id=obs_id,
        sheet=_SHEET_META,
        row=meta_row_num,
        path=index.path,
    )

    z_arr = index.z_records[key]
    x_arr = index.x_records[key]

    # Per-channel arrays are indexed in total-channel space. A survey
    # may record several components per physical sensor (e.g. amenomiya:
    # 24 sensors x 2 components Y+Z = 48 channels), so the array length
    # is num_sensor * n_components (n_components >= 1). A length that is
    # not a positive multiple of num_sensor, or z/x lengths that
    # disagree, indicates a genuine data-entry error.
    if len(z_arr) == 0 or len(z_arr) % num_sensor != 0:
        raise ValueError(
            f"Observation id '{obs_id}': sheet '{_SHEET_Z}' row "
            f"{z_row_num} provides {len(z_arr)} non-null channel "
            f"values, which is not a positive multiple of "
            f"'{_HDR_NUM_SENSOR}'={num_sensor} in sheet '{_SHEET_META}' "
            f"row {meta_row_num} of datasheet '{index.path}'."
        )
    if len(x_arr) != len(z_arr):
        raise ValueError(
            f"Observation id '{obs_id}': sheet '{_SHEET_X}' row "
            f"{x_row_num} provides {len(x_arr)} non-null channel "
            f"values, but sheet '{_SHEET_Z}' row {z_row_num} provides "
            f"{len(z_arr)}; the per-channel arrays must share the same "
            f"total-channel length in datasheet '{index.path}'."
        )

    total_channels = len(z_arr)

    # Optional 'distance' sheet override. Absence (sheet missing OR
    # this obs_id missing from the sheet) is silent and yields None;
    # presence is validated for per-channel length, matching the
    # z_distance / x_distance contract.
    distance_arr: Optional[Tuple[float, ...]]
    if key in index.distance_records:
        distance_arr = index.distance_records[key]
        distance_row_num = index.distance_row_numbers[key]
        if len(distance_arr) != total_channels:
            raise ValueError(
                f"Observation id '{obs_id}': sheet '{_SHEET_DISTANCE}' "
                f"row {distance_row_num} provides {len(distance_arr)} "
                f"non-null channel values, but the per-channel arrays "
                f"have total-channel length {total_channels} in "
                f"datasheet '{index.path}'."
            )
    else:
        distance_arr = None

    return NormalizedMetadata(
        obs_id=meta.get(_HDR_NO),
        shot_method=meta.get(_HDR_SHOT_METHOD),
        condition=meta.get(_HDR_CONDITION),
        excdistance=meta.get(_HDR_EXCDISTANCE),
        source_x=_coerce_float(
            meta.get(_HDR_SOURCE_X),
            field=_HDR_SOURCE_X,
            obs_id=obs_id,
            sheet=_SHEET_META,
            row=meta_row_num,
            path=index.path,
        ),
        sensor1_x=_coerce_float(
            meta.get(_HDR_SENSOR1_X),
            field=_HDR_SENSOR1_X,
            obs_id=obs_id,
            sheet=_SHEET_META,
            row=meta_row_num,
            path=index.path,
        ),
        interval=_coerce_float(
            meta.get(_HDR_INTERVAL),
            field=_HDR_INTERVAL,
            obs_id=obs_id,
            sheet=_SHEET_META,
            row=meta_row_num,
            path=index.path,
        ),
        num_sensor=num_sensor,
        x_chs_range=meta.get(_HDR_X_CHS),
        y_chs_range=meta.get(_HDR_Y_CHS),
        z_chs_range=meta.get(_HDR_Z_CHS),
        x_reverse=meta.get(_HDR_X_REVERSE),
        y_reverse=meta.get(_HDR_Y_REVERSE),
        z_reverse=meta.get(_HDR_Z_REVERSE),
        z_distance=z_arr,
        x_distance=x_arr,
        distance=distance_arr,
    )


# ---------------------------------------------------------------------------
# Internal helpers.
# ---------------------------------------------------------------------------


def _obs_key(value: Any) -> Optional[str]:
    """Normalize an observation id to a canonical string lookup key.

    - ``None`` and empty strings -> ``None`` (treated as "no key").
    - Plain integers and integer-valued floats -> their decimal string
      (so ``891`` and ``891.0`` and ``"891"`` all map to ``"891"``).
    - Strings -> stripped string.
    - Anything else -> ``str(value)``.
    """
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    if isinstance(value, bool):
        # bool is a subclass of int; keep an explicit branch for clarity.
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def _index_meta_sheet(
    ws: Any, path: Path, required_headers: Tuple[str, ...]
) -> Tuple[Dict[str, Mapping[str, Any]], Dict[str, int]]:
    """Scan ``Sheet1`` once and return (records, row_numbers) maps.

    ``records[obs_key]`` is a dict keyed by header name with the raw
    cell value for each Sheet1 column. ``row_numbers[obs_key]`` is the
    1-based row index in the source workbook (header row is row 1).
    """
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError(
            f"Sheet '{ws.title}' in datasheet '{path}' is empty."
        )

    header_to_idx = _build_header_index(ws.title, header, path)
    _check_required_headers(ws.title, header_to_idx, required_headers, path)

    no_idx = header_to_idx[_HDR_NO]
    records: Dict[str, Mapping[str, Any]] = {}
    row_numbers: Dict[str, int] = {}

    for row_num, row in enumerate(rows, start=2):
        if no_idx >= len(row):
            continue
        raw_no = row[no_idx]
        key = _obs_key(raw_no)
        if key is None:
            continue
        if key in records:
            raise ValueError(
                f"Sheet '{ws.title}' in datasheet '{path}' has a "
                f"duplicate observation id {raw_no!r} at row {row_num} "
                f"(already seen at row {row_numbers[key]})."
            )
        record: Dict[str, Any] = {}
        for header_name, idx in header_to_idx.items():
            record[header_name] = row[idx] if idx < len(row) else None
        records[key] = record
        row_numbers[key] = row_num

    return records, row_numbers


def _index_dist_sheet(
    ws: Any, path: Path, required_headers: Tuple[str, ...]
) -> Tuple[Dict[str, Tuple[float, ...]], Dict[str, int]]:
    """Scan a distance sheet once and return (channel_arrays, row_numbers).

    Channels are read as the contiguous run ``ch1``, ``ch2``, ... and
    truncated at the first column whose value is ``None``. The returned
    tuple length therefore reflects the actual non-null run; the
    ``num_sensor`` cross-check happens later in :func:`get_record`.
    """
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError(
            f"Sheet '{ws.title}' in datasheet '{path}' is empty."
        )

    header_to_idx = _build_header_index(ws.title, header, path)
    _check_required_headers(ws.title, header_to_idx, required_headers, path)

    # Collect ch1, ch2, ... in declared order, stopping at the first
    # missing channel header (so a workbook that declares ch1..ch96
    # contributes 96 column indices regardless of where they sit).
    ch_indices = []
    n = 1
    while True:
        header_name = f"{_DIST_CHANNEL_PREFIX}{n}"
        idx = header_to_idx.get(header_name)
        if idx is None:
            break
        ch_indices.append(idx)
        n += 1
    if not ch_indices:
        raise ValueError(
            f"Sheet '{ws.title}' in datasheet '{path}' has no '"
            f"{_DIST_CHANNEL_PREFIX}1' column; expected at least one "
            f"per-channel column starting at '{_DIST_CHANNEL_PREFIX}1'."
        )

    no_idx = header_to_idx[_HDR_NO]
    arrays: Dict[str, Tuple[float, ...]] = {}
    row_numbers: Dict[str, int] = {}

    for row_num, row in enumerate(rows, start=2):
        if no_idx >= len(row):
            continue
        raw_no = row[no_idx]
        key = _obs_key(raw_no)
        if key is None:
            continue
        if key in arrays:
            raise ValueError(
                f"Sheet '{ws.title}' in datasheet '{path}' has a "
                f"duplicate observation id {raw_no!r} at row {row_num} "
                f"(already seen at row {row_numbers[key]})."
            )

        values = []
        for ci in ch_indices:
            if ci >= len(row):
                break
            v = row[ci]
            if v is None:
                break
            try:
                values.append(float(v))
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f"Sheet '{ws.title}' row {row_num} (observation id "
                    f"{raw_no!r}) in datasheet '{path}': channel value "
                    f"{v!r} is not convertible to float ({exc})."
                ) from exc
        arrays[key] = tuple(values)
        row_numbers[key] = row_num

    return arrays, row_numbers


def _build_header_index(
    sheet_name: str, header_row: Tuple[Any, ...], path: Path
) -> Dict[str, int]:
    """Build a {header_name: column_index_0_based} map from row 1."""
    header_to_idx: Dict[str, int] = {}
    for i, value in enumerate(header_row):
        if value is None:
            continue
        if value in header_to_idx:
            raise ValueError(
                f"Sheet '{sheet_name}' in datasheet '{path}' has a "
                f"duplicate header {value!r} at column {i + 1} "
                f"(already at column {header_to_idx[value] + 1})."
            )
        header_to_idx[value] = i
    return header_to_idx


def _check_required_headers(
    sheet_name: str,
    header_to_idx: Mapping[str, int],
    required: Tuple[str, ...],
    path: Path,
) -> None:
    missing = [h for h in required if h not in header_to_idx]
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' in datasheet '{path}' is missing "
            f"required header(s): {missing}. Found headers: "
            f"{list(header_to_idx)[:20]}{'...' if len(header_to_idx) > 20 else ''}."
        )


def _coerce_num_sensor(
    value: Any, *, obs_id: Any, sheet: str, row: int, path: Path
) -> int:
    if value is None:
        raise ValueError(
            f"Observation id '{obs_id}': '{_HDR_NUM_SENSOR}' is empty "
            f"in sheet '{sheet}' row {row} of datasheet '{path}'."
        )
    try:
        coerced = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Observation id '{obs_id}': '{_HDR_NUM_SENSOR}' must be "
            f"an integer, got {value!r} in sheet '{sheet}' row {row} "
            f"of datasheet '{path}' ({exc})."
        ) from exc
    if coerced <= 0:
        raise ValueError(
            f"Observation id '{obs_id}': '{_HDR_NUM_SENSOR}' must be "
            f"a positive integer, got {coerced} in sheet '{sheet}' "
            f"row {row} of datasheet '{path}'."
        )
    return coerced


def _coerce_float(
    value: Any, *, field: str, obs_id: Any, sheet: str, row: int, path: Path
) -> float:
    if value is None:
        raise ValueError(
            f"Observation id '{obs_id}': '{field}' is empty in sheet "
            f"'{sheet}' row {row} of datasheet '{path}'."
        )
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Observation id '{obs_id}': '{field}' must be numeric, "
            f"got {value!r} in sheet '{sheet}' row {row} of datasheet "
            f"'{path}' ({exc})."
        ) from exc


__all__ = [
    "DatasheetIndex",
    "NormalizedMetadata",
    "load_datasheet",
    "get_record",
]
